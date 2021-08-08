from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from os import path
import smtplib

from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, path_file):
        self.path_file = path_file
        self.workbook = self.load_file()
        self.n_column = self.count_columns()
        self.n_rows = self.count_rows()

    def load_file(self):
        return load_workbook(self.path_file).active

    def count_rows(self):
        for nrow, cell in enumerate(self.workbook['A']):
            if cell.value is None:
                return nrow

    def count_columns(self):
        return self.workbook.max_column

    def extract_data(self, index):
        excel_data = []
        for row in self.workbook[index]:
            excel_data.append(row.value)
        return excel_data[:2], excel_data[2:]


class Email:
    def __init__(self, transmitter, passwd, contact_info):
        self.transmitter = transmitter
        self.password = passwd
        self.receiver = contact_info[0]
        self.receiver_name = contact_info[1]

    def create_message(self, data):
        message = MIMEMultipart()
        message['From'] = self.transmitter
        message['To'] = self.receiver
        message['Subject'] = 'Notas [SEMESTRE-PERIODO, etc]'

        html_body = f'''
                    <html>
         	            <head> </head>
         	            <body>
                            Good morning, {self.receiver_name}

                            <p> Below are the programming course notes. The global grade is divided 				
                            as follows: Examen (50%) y Workshop (50%). <br> <br>

                            In general, your notes are: <br>
                            <ul>
                                <li type="disc"><b>Examen</b>: {data[1]}</li>
                                <li type="disc"><b>Workshop</b>: {data[2]}</li>
                            </ul>

                            The total grade for the course is: {data[0]} <br> <br>
                            <br> <br>
                            </p
                        </body>
                    </html>
                    '''
        message.attach(MIMEText(html_body, 'html'))
        return message.as_string()

    def send_mail(self, text):
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(self.transmitter, self.password)
        server.sendmail(self.transmitter, self.receiver, text)
        server.quit()


if __name__ == '__main__':
    email_address = input('Enter your email: ')
    passwd = input('Enter your password: ')
    file_path = input('Enter file path: ')

    if path.isfile(file_path):
        file = ExcelReader(file_path)
        for i in range(1, file.n_rows + 1):
            contact_info, data = file.extract_data(i)
            email = Email(email_address, passwd, contact_info)
            email.send_mail(email.create_message(data))
            print('Email send to ', contact_info[1])
    else:
        print("Error, the file does not exists")
